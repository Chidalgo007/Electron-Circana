import os
from openpyxl import load_workbook
import datetime
import sys
import io

# === UTF-8 output for Electron logs ===
sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
sys.stderr = io.TextIOWrapper(sys.stderr.buffer, encoding='utf-8', errors='replace')

# === Helper functions ===
def force_console_output(message):
    """Print and flush immediately (so Electron sees it)."""
    try:
        print(message)
    except UnicodeEncodeError:
        print(message.encode("utf-8", errors="replace").decode("utf-8"))
    sys.stdout.flush()
    sys.stderr.flush()

# ========== Main Automation Logic ========== #
def run_excel_updates(destination_folder):
    try:
        # === Calendar Sheet Update ===
        force_console_output("\n\n=== Updating Calendar Sheet ===")
        calendar_file_path = os.path.join(destination_folder, "Calendar.xlsx")
        if not os.path.exists(calendar_file_path):
            raise FileNotFoundError(f"❌  File not found: {calendar_file_path}")

        wb = load_workbook(calendar_file_path)
        sheet = wb.active
        
        today = datetime.date.today()
        sunday = today - datetime.timedelta(days=today.weekday() + 8) # 2 week's back Sunday

        sheet["J1"].value = sunday
        force_console_output(f"📅  Calendar.xlsx: J1 updated to {sunday}")
        wb.save(calendar_file_path)

        # === Week Sheet Update ===
        force_console_output("\n\n=== Updating Week Sheet ===")
        week_file_path = os.path.join(destination_folder, "Weeks.xlsx")
        
        if not os.path.exists(week_file_path):
            raise FileNotFoundError(f"❌  File not found: {week_file_path}")
        wb = load_workbook(week_file_path)
        sheet = wb.active

        sheet["A2"].value = sunday
        force_console_output(f"📅  Weeks.xlsx: A2 updated to {sunday}")
        wb.save(week_file_path)

        return True
    except Exception as e:
        force_console_output(f"❌  Error occurred: {e}")
        return False
        

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(1)

    file_path = sys.argv[1]
    success = run_excel_updates(file_path)
    sys.exit(0 if success else 1)
